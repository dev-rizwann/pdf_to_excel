import re
from pathlib import Path
from datetime import datetime, date
import os

import pdfplumber
import pandas as pd
from openpyxl.utils import get_column_letter

# ---------- regex ----------
DATE_RE = re.compile(r"^\d{4}/\d{1,2}/\d{1,2}$")
TXN_RE = re.compile(r"^\d{4}$")  # 4-digit transaction (your "Order #")

# ✅ FIX #1: allow long order ids that sometimes come like 7315111772497_2
ORDER_LONG_RE = re.compile(r"^\d{10,}(?:_\d+)?$")

NUM_RE = re.compile(r"^\d+(?:\.\d{1,2})?$")

# invoice total patterns (based on your PDF)
TOTAL_USD_RE = re.compile(
    r"Total\s*\(\s*USD\s*\)\s*:\s*([$]?\s*[\d,]+(?:\.\d+)?)", re.IGNORECASE
)
TOTAL_AMOUNT_RE = re.compile(
    r"TOTAL\s+AMOUNT\s*[$]?\s*([\d,]+(?:\.\d+)?)", re.IGNORECASE
)

SINGLE_COUNTRIES = {
    "canada", "australia", "mexico", "china", "france", "germany", "italy",
    "spain", "japan", "singapore", "uae", "pakistan", "uk"
}

# ---------- helpers ----------
def clean(t: str) -> str:
    return (t or "").strip().strip("[]{}(),;:")

def pick_files_dialog():
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    paths = filedialog.askopenfilenames(
        title="Select invoice PDFs",
        filetypes=[("PDF files", "*.pdf")]
    )
    return [Path(p) for p in paths]

def extract_tokens_and_text(pdf_path: Path):
    """
    Returns (tokens, full_text).
    tokens => used for row parsing
    full_text => used for invoice totals extraction
    """
    tokens = []
    full_text_parts = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            txt = page.extract_text(x_tolerance=2, y_tolerance=2) or ""
            if txt:
                full_text_parts.append(txt)
                tokens.extend([t for t in re.split(r"\s+", txt) if t])
    return tokens, "\n".join(full_text_parts)

def extract_total_usd(full_text: str):
    """
    Prefer: Total(USD): <number>
    Fallback: TOTAL AMOUNT <number>
    Returns float or None.
    """
    if not full_text:
        return None

    m = TOTAL_USD_RE.search(full_text)
    if not m:
        m = TOTAL_AMOUNT_RE.search(full_text)
    if not m:
        return None

    raw = (m.group(1) or "").strip()
    raw = raw.replace("$", "").replace(" ", "").replace(",", "")
    try:
        return float(raw)
    except:
        return None

def stop_at_total_usd(tokens):
    # Prevent totals section from being mis-read as rows/prices
    for i, tok in enumerate(tokens):
        if clean(tok).lower().startswith("total(usd"):
            return tokens[:i]
    return tokens

def parse_ymd_to_date(s: str) -> date:
    y, m, d = s.split("/")
    return date(int(y), int(m), int(d))

def excel_serial_from_date(dt: date) -> int:
    excel_epoch = date(1899, 12, 30)  # Windows Excel epoch
    return (dt - excel_epoch).days

def is_country_at(tokens, i):
    t = clean(tokens[i]).lower()
    if t == "united" and i + 1 < len(tokens) and clean(tokens[i + 1]).lower() == "states":
        return True, i + 1
    if t in SINGLE_COUNTRIES:
        return True, i
    return False, -1

def is_row_start(tokens, i):
    # Anchor: long order id + 4-digit txn + date
    if i + 2 >= len(tokens):
        return False
    return (
        ORDER_LONG_RE.match(clean(tokens[i])) and
        TXN_RE.match(clean(tokens[i + 1])) and
        DATE_RE.match(clean(tokens[i + 2]))
    )

def qty_before_country(tokens, row_start, country_pos):
    # last integer 1..999 before country within this row
    for j in range(country_pos - 1, row_start, -1):
        t = clean(tokens[j])
        if t.isdigit():
            v = int(t)
            if 1 <= v <= 999:
                return v
    return None

def looks_like_price(tok: str):
    t = clean(tok)

    # ✅ FIX #2: never treat 4-digit txn values as prices
    if TXN_RE.match(t):
        return False

    if not NUM_RE.match(t):
        return False

    # Don't treat long order ids as prices
    if ORDER_LONG_RE.match(t):
        return False

    try:
        v = float(t)
    except:
        return False

    # reject noise like "1" / "2"
    if "." not in t and v < 10:
        return False

    if v <= 0 or v > 100000:
        return False

    return True

def last_price_after_country(tokens, start_after_country, row_end):
    last = None
    for j in range(start_after_country, row_end):
        if looks_like_price(tokens[j]):
            last = float(clean(tokens[j]))
    return last

def to_dutch_text(num):
    if num is None:
        return ""
    return f"{float(num):.2f}".replace(".", ",")

def parse_pdf_tokens(tokens, file_name):
    tokens = stop_at_total_usd(tokens)

    rows = []
    i = 0
    n = len(tokens)

    while i < n:
        if not is_row_start(tokens, i):
            i += 1
            continue

        txn_s = clean(tokens[i + 1])
        date_s = clean(tokens[i + 2])

        order4 = int(txn_s) if TXN_RE.match(txn_s) else None
        dt = parse_ymd_to_date(date_s) if DATE_RE.match(date_s) else None
        date_serial = excel_serial_from_date(dt) if dt else None

        # find row end at next row start
        row_end = min(n, i + 320)
        j = i + 3
        while j < row_end:
            if is_row_start(tokens, j):
                row_end = j
                break
            j += 1

        # find first country
        country_pos = -1
        country_end = -1
        k = i + 3
        while k < row_end:
            ok, cend = is_country_at(tokens, k)
            if ok:
                country_pos = k
                country_end = cend
                break
            k += 1

        qty = None
        cost = None
        if country_pos != -1:
            qty = qty_before_country(tokens, i, country_pos)
            cost = last_price_after_country(tokens, country_end + 1, row_end)

        # Only keep real rows
        if order4 is not None and date_serial is not None and cost is not None:
            rows.append({
                "File Name": file_name,
                "DateSerial": date_serial,     # e.g. 46033
                "Order #": order4,             # 4-digit txn
                "Qty": int(qty) if qty is not None else None,
                "Cost": float(cost),           # numeric
                "Cost_NL": to_dutch_text(cost) # dutch text
            })

        i = row_end

    return rows

def autosize_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

def apply_formats(writer, sheet_name="COGS"):
    ws = writer.sheets[sheet_name]
    headers = {str(c.value).strip(): c.column for c in ws[1]}
    last_row = ws.max_row

    col_date = headers.get("DateSerial")
    col_ord  = headers.get("Order #")
    col_qty  = headers.get("Qty")
    col_cost = headers.get("Cost")
    col_nl   = headers.get("Cost_NL")

    if col_date:
        for r in range(2, last_row + 1):
            ws.cell(r, col_date).number_format = "0"

    if col_ord:
        for r in range(2, last_row + 1):
            ws.cell(r, col_ord).number_format = "0"

    if col_qty:
        for r in range(2, last_row + 1):
            ws.cell(r, col_qty).number_format = "0"

    if col_cost:
        for r in range(2, last_row + 1):
            ws.cell(r, col_cost).number_format = "#,##0.00"

    if col_nl:
        for r in range(2, last_row + 1):
            ws.cell(r, col_nl).number_format = "@"  # text

    autosize_columns(ws)

def apply_formats_totals(writer, sheet_name="InvoiceTotals"):
    ws = writer.sheets[sheet_name]
    headers = {str(c.value).strip(): c.column for c in ws[1]}
    last_row = ws.max_row

    col_total = headers.get("Total_USD_Extracted")
    col_sum   = headers.get("COGS_Sum_Formula")
    col_diff  = headers.get("Diff")
    col_flag  = headers.get("Match")

    if col_total:
        for r in range(2, last_row + 1):
            ws.cell(r, col_total).number_format = "#,##0.00"

    if col_sum:
        for r in range(2, last_row + 1):
            ws.cell(r, col_sum).number_format = "#,##0.00"

    if col_diff:
        for r in range(2, last_row + 1):
            ws.cell(r, col_diff).number_format = "#,##0.00"

    if col_flag:
        for r in range(2, last_row + 1):
            ws.cell(r, col_flag).number_format = "@"

    autosize_columns(ws)

def open_file_in_excel(path: Path):
    os.startfile(str(path))  # Windows only

def main():
    pdfs = pick_files_dialog()
    if not pdfs:
        return

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_xlsx = Path.cwd() / f"COGS_{ts}.xlsx"

    all_rows = []
    totals_rows = []
    log_rows = []

    for pdf in pdfs:
        try:
            tokens, full_text = extract_tokens_and_text(pdf)

            # row-level extraction
            rows = parse_pdf_tokens(tokens, pdf.name)
            all_rows.extend(rows)

            # totals extraction
            total_usd = extract_total_usd(full_text)
            totals_rows.append({
                "File Name": pdf.name,
                "Total_USD_Extracted": total_usd
            })

            log_rows.append({
                "File": pdf.name,
                "Tokens": len(tokens),
                "Rows": len(rows),
                "Status": "OK",
                "Error": ""
            })
        except Exception as e:
            totals_rows.append({
                "File Name": pdf.name,
                "Total_USD_Extracted": None
            })
            log_rows.append({
                "File": pdf.name,
                "Tokens": "",
                "Rows": "",
                "Status": "ERROR",
                "Error": str(e)
            })

    df = pd.DataFrame(all_rows, columns=["File Name", "DateSerial", "Order #", "Qty", "Cost", "Cost_NL"])
    dflog = pd.DataFrame(log_rows, columns=["File", "Tokens", "Rows", "Status", "Error"])

    # Totals sheet (formulas injected after write)
    dftot = pd.DataFrame(totals_rows, columns=["File Name", "Total_USD_Extracted"])
    dftot["COGS_Sum_Formula"] = ""
    dftot["Diff"] = ""
    dftot["Match"] = ""

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="COGS")
        dftot.to_excel(writer, index=False, sheet_name="InvoiceTotals")
        dflog.to_excel(writer, index=False, sheet_name="Log")

        apply_formats(writer, sheet_name="COGS")
        apply_formats_totals(writer, sheet_name="InvoiceTotals")

        ws_log = writer.sheets["Log"]
        autosize_columns(ws_log)

        # ----- Inject formulas into InvoiceTotals -----
        ws_tot = writer.sheets["InvoiceTotals"]
        tot_headers = {str(c.value).strip(): c.column for c in ws_tot[1]}

        col_file  = tot_headers["File Name"]              # A
        col_total = tot_headers["Total_USD_Extracted"]     # B
        col_sum   = tot_headers["COGS_Sum_Formula"]        # C
        col_diff  = tot_headers["Diff"]                    # D
        col_match = tot_headers["Match"]                   # E

        last_row = ws_tot.max_row
        for r in range(2, last_row + 1):
            file_cell  = ws_tot.cell(r, col_file).coordinate
            total_cell = ws_tot.cell(r, col_total).coordinate
            sum_cell   = ws_tot.cell(r, col_sum).coordinate
            diff_cell  = ws_tot.cell(r, col_diff).coordinate

            # COGS!A = File Name, COGS!E = Cost
            ws_tot.cell(r, col_sum).value = f'=SUMIF(COGS!$A:$A,{file_cell},COGS!$E:$E)'
            ws_tot.cell(r, col_diff).value = f'={sum_cell}-{total_cell}'
            ws_tot.cell(r, col_match).value = f'=IF(ABS({diff_cell})<0.01,"OK","CHECK")'

    open_file_in_excel(out_xlsx)

if __name__ == "__main__":
    main()
