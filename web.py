# converter.py

from pathlib import Path
from datetime import datetime, date
import re
import pdfplumber
import pandas as pd
from openpyxl.utils import get_column_letter

# =====================
# Regex definitions
# =====================
DATE_RE = re.compile(r"^\d{4}/\d{1,2}/\d{1,2}$")
TXN_RE = re.compile(r"^\d{4}$")
ORDER_LONG_RE = re.compile(r"^\d{10,}(?:_\d+)?$")
NUM_RE = re.compile(r"^\d+(?:\.\d{1,2})?$")

TOTAL_USD_RE = re.compile(
    r"Total\s*\(\s*USD\s*\)\s*:\s*([$]?\s*[\d,]+(?:\.\d+)?)",
    re.IGNORECASE,
)

TOTAL_AMOUNT_RE = re.compile(
    r"TOTAL\s+AMOUNT\s*[$]?\s*([\d,]+(?:\.\d+)?)",
    re.IGNORECASE,
)

SINGLE_COUNTRIES = {
    "canada", "australia", "mexico", "china", "france", "germany",
    "italy", "spain", "japan", "singapore", "uae", "pakistan", "uk"
}

# =====================
# Helper functions
# =====================
def clean(t: str) -> str:
    return (t or "").strip().strip("[]{}(),;:")

def parse_ymd_to_date(s: str) -> date:
    y, m, d = s.split("/")
    return date(int(y), int(m), int(d))

def excel_serial_from_date(dt: date) -> int:
    excel_epoch = date(1899, 12, 30)
    return (dt - excel_epoch).days

def extract_tokens_and_text(pdf_path: Path):
    tokens = []
    full_text_parts = []

    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            txt = page.extract_text(x_tolerance=2, y_tolerance=2) or ""
            if txt:
                full_text_parts.append(txt)
                tokens.extend(re.split(r"\s+", txt))

    return tokens, "\n".join(full_text_parts)

def extract_total_usd(full_text: str):
    if not full_text:
        return None

    m = TOTAL_USD_RE.search(full_text) or TOTAL_AMOUNT_RE.search(full_text)
    if not m:
        return None

    raw = (
        m.group(1)
        .replace("$", "")
        .replace(" ", "")
        .replace(",", "")
    )

    try:
        return float(raw)
    except ValueError:
        return None

def is_country_at(tokens, i):
    t = clean(tokens[i]).lower()

    if t == "united" and i + 1 < len(tokens):
        if clean(tokens[i + 1]).lower() == "states":
            return True, i + 1

    if t in SINGLE_COUNTRIES:
        return True, i

    return False, -1

def is_row_start(tokens, i):
    if i + 2 >= len(tokens):
        return False

    return (
        ORDER_LONG_RE.match(clean(tokens[i]))
        and TXN_RE.match(clean(tokens[i + 1]))
        and DATE_RE.match(clean(tokens[i + 2]))
    )

def qty_before_country(tokens, row_start, country_pos):
    for j in range(country_pos - 1, row_start, -1):
        t = clean(tokens[j])
        if t.isdigit():
            v = int(t)
            if 1 <= v <= 999:
                return v
    return None

def looks_like_price(tok: str):
    t = clean(tok)

    if TXN_RE.match(t):
        return False
    if not NUM_RE.match(t):
        return False
    if ORDER_LONG_RE.match(t):
        return False

    try:
        v = float(t)
    except ValueError:
        return False

    if "." not in t and v < 10:
        return False
    if v <= 0 or v > 100000:
        return False

    return True

def last_price_after_country(tokens, start, row_end):
    last = None
    for j in range(start, row_end):
        if looks_like_price(tokens[j]):
            last = float(clean(tokens[j]))
    return last

def stop_at_total_usd(tokens):
    for i, tok in enumerate(tokens):
        if clean(tok).lower().startswith("total(usd"):
            return tokens[:i]
    return tokens

def to_dutch_text(num):
    if num is None:
        return ""
    return f"{float(num):.2f}".replace(".", ",")

# =====================
# Core parsing logic
# =====================
def parse_pdf_tokens(tokens, file_name):
    tokens = stop_at_total_usd(tokens)
    rows = []
    i = 0
    n = len(tokens)

    while i < n:
        if not is_row_start(tokens, i):
            i += 1
            continue

        txn_s = clean(tokens[i + 1])
        date_s = clean(tokens[i + 2])

        order4 = int(txn_s) if TXN_RE.match(txn_s) else None
        dt = parse_ymd_to_date(date_s) if DATE_RE.match(date_s) else None
        date_serial = excel_serial_from_date(dt) if dt else None

        row_end = min(n, i + 320)
        j = i + 3
        while j < row_end:
            if is_row_start(tokens, j):
                row_end = j
                break
            j += 1

        country_pos = -1
        country_end = -1
        k = i + 3
        while k < row_end:
            ok, cend = is_country_at(tokens, k)
            if ok:
                country_pos = k
                country_end = cend
                break
            k += 1

        qty = None
        cost = None

        if country_pos != -1:
            qty = qty_before_country(tokens, i, country_pos)
            cost = last_price_after_country(tokens, country_end + 1, row_end)

        if order4 and date_serial and cost is not None:
            rows.append({
                "File Name": file_name,
                "DateSerial": date_serial,
                "Order #": order4,
                "Qty": int(qty) if qty else None,
                "Cost": float(cost),
                "Cost_NL": to_dutch_text(cost),
            })

        i = row_end

    return rows

# =====================
# Excel helpers
# =====================
def autosize_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

def apply_formats(writer, sheet_name="COGS"):
    ws = writer.sheets[sheet_name]
    headers = {str(c.value).strip(): c.column for c in ws[1]}
    last_row = ws.max_row

    for r in range(2, last_row + 1):
        if "DateSerial" in headers:
            ws.cell(r, headers["DateSerial"]).number_format = "0"
        if "Order #" in headers:
            ws.cell(r, headers["Order #"]).number_format = "0"
        if "Qty" in headers:
            ws.cell(r, headers["Qty"]).number_format = "0"
        if "Cost" in headers:
            ws.cell(r, headers["Cost"]).number_format = "#,##0.00"
        if "Cost_NL" in headers:
            ws.cell(r, headers["Cost_NL"]).number_format = "@"

    autosize_columns(ws)

def apply_formats_totals(writer, sheet_name="InvoiceTotals"):
    ws = writer.sheets[sheet_name]
    headers = {str(c.value).strip(): c.column for c in ws[1]}
    last_row = ws.max_row

    for r in range(2, last_row + 1):
        if "Total_USD_Extracted" in headers:
            ws.cell(r, headers["Total_USD_Extracted"]).number_format = "#,##0.00"
        if "COGS_Sum_Formula" in headers:
            ws.cell(r, headers["COGS_Sum_Formula"]).number_format = "#,##0.00"
        if "Diff" in headers:
            ws.cell(r, headers["Diff"]).number_format = "#,##0.00"
        if "Match" in headers:
            ws.cell(r, headers["Match"]).number_format = "@"

    autosize_columns(ws)

# =====================
# Main converter
# =====================
def convert_pdfs_to_excel(pdf_paths, output_dir):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_xlsx = output_dir / f"COGS_{ts}.xlsx"

    all_rows = []
    totals_rows = []
    log_rows = []

    for pdf in pdf_paths:
        try:
            tokens, full_text = extract_tokens_and_text(pdf)
            rows = parse_pdf_tokens(tokens, pdf.name)
            all_rows.extend(rows)

            total_usd = extract_total_usd(full_text)
            totals_rows.append({
                "File Name": pdf.name,
                "Total_USD_Extracted": total_usd,
            })

            log_rows.append({
                "File": pdf.name,
                "Tokens": len(tokens),
                "Rows": len(rows),
                "Status": "OK",
                "Error": "",
            })

        except Exception as e:
            print(f"[ERROR] Failed to process {pdf.name}: {e}")
            totals_rows.append({
                "File Name": pdf.name,
                "Total_USD_Extracted": None,
            })
            log_rows.append({
                "File": pdf.name,
                "Tokens": "",
                "Rows": "",
                "Status": "ERROR",
                "Error": str(e),
            })

    if not all_rows and not totals_rows:
        print("[WARNING] No rows extracted.")
        return None

    df = pd.DataFrame(all_rows)
    dftot = pd.DataFrame(totals_rows)
    dflog = pd.DataFrame(log_rows)

    dftot["COGS_Sum_Formula"] = ""
    dftot["Diff"] = ""
    dftot["Match"] = ""

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="COGS")
        dftot.to_excel(writer, index=False, sheet_name="InvoiceTotals")
        dflog.to_excel(writer, index=False, sheet_name="Log")

        apply_formats(writer, "COGS")
        apply_formats_totals(writer, "InvoiceTotals")
        autosize_columns(writer.sheets["Log"])

        ws = writer.sheets["InvoiceTotals"]
        headers = {c.value: c.column for c in ws[1]}

        for r in range(2, ws.max_row + 1):
            ws.cell(r, headers["COGS_Sum_Formula"]).value = (
                f'=SUMIF(COGS!$A:$A,A{r},COGS!$E:$E)'
            )
            ws.cell(r, headers["Diff"]).value = f'=C{r}-B{r}'
            ws.cell(r, headers["Match"]).value = (
                f'=IF(ABS(D{r})<0.01,"OK","CHECK")'
            )

    print(f"[INFO] Excel file created: {out_xlsx}")
    return out_xlsx
#love 
